from __future__ import annotations

import argparse
from collections import Counter
import hashlib
import json
from pathlib import Path
import sys
import zipfile

import numpy as np
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parent
sys.path.insert(0, str(ROOT / "src"))

from port_queue.config import load_config  # noqa: E402
from port_queue.public_data import (  # noqa: E402
    clean_sc_ports,
    load_sc_ports,
    make_public_data_scenarios,
    parse_houston_reports,
    parse_virginia_metrics,
    summarize_capacity_benchmark,
    summarize_dwelling_time,
    summarize_sc_ports,
    summarize_tas_benchmark,
)


OUTPUT_FILES = (
    "clean_port_houston_terminal_reports.csv",
    "clean_port_virginia_weekly_metrics.csv",
    "dwelling_activity_counts.csv",
    "mendeley_capacity_benchmark_summary.csv",
    "mendeley_tas_benchmark_summary.csv",
    "public_data_calibrated_scenarios.csv",
    "sc_ports_daily_gate_missions.csv",
    "sc_ports_summary.csv",
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def extract_mendeley(raw: Path, target: Path) -> dict[str, Path]:
    groups = {
        "mendeley_tas_tours": raw / "mendeley_tas_tours" / "82zzdkrxx8-v1.zip",
        "mendeley_capacity_management": raw / "mendeley_capacity_management" / "2b646hgkt7-v1.zip",
        "mendeley_dwelling_time": raw / "mendeley_dwelling_time" / "yvp2b4rtp3-v1.zip",
    }
    extracted: dict[str, Path] = {}
    for group, archive in groups.items():
        destination = target / group
        destination.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(archive) as handle:
            bad = handle.testzip()
            if bad:
                raise RuntimeError(f"corrupt member in {archive}: {bad}")
            handle.extractall(destination)
        extracted[group] = destination
    return extracted


def run_transform(
    *,
    sc_ports: Path,
    virginia: Path,
    houston: Path,
    tas: Path,
    capacity: Path,
    dwelling: Path,
    output: Path,
) -> pd.DataFrame:
    output.mkdir(parents=True, exist_ok=True)
    config = load_config(ROOT / "configs" / "fidelity.json")
    frames = load_sc_ports(sc_ports)
    cleaned = clean_sc_ports(frames, output / "clean_sc_ports")
    sc_summary, daily = summarize_sc_ports(cleaned)
    sc_summary.to_csv(output / "sc_ports_summary.csv", index=False, encoding="utf-8-sig")
    daily.to_csv(output / "sc_ports_daily_gate_missions.csv", index=False, encoding="utf-8-sig")
    va = parse_virginia_metrics(virginia, output)
    tx = parse_houston_reports(houston, output)
    tas_files = sorted(tas.rglob("TAS_experiments_input_data.xlsx"))
    if not tas_files:
        raise FileNotFoundError(f"TAS_experiments_input_data.xlsx not found under {tas}")
    summarize_tas_benchmark(tas_files[0], output)
    summarize_capacity_benchmark(capacity, output)
    _summarize_dwelling_compatible(dwelling, output)
    scenarios, _ = make_public_data_scenarios(config, daily, va, tx)
    scenarios.to_csv(output / "public_data_calibrated_scenarios.csv", index=False, encoding="utf-8-sig")
    return scenarios


def _summarize_dwelling_compatible(dwelling: Path, output: Path) -> None:
    from openpyxl import load_workbook

    files = sorted(dwelling.rglob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"dwelling workbook not found under {dwelling}")
    workbook = load_workbook(files[0], read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    workbook.close()
    if "Start Timestamp" in header and "End Timestamp" in header:
        summarize_dwelling_time(dwelling, output)
        return
    if "Activity" not in header:
        raise KeyError(f"unsupported dwelling columns: {header}")

    workbook = load_workbook(files[0], read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    activity_index = header.index("Activity")
    counts: Counter[str] = Counter()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        value = row[activity_index]
        if value is not None:
            counts[str(value)] += 1
    workbook.close()
    frame = pd.DataFrame(counts.most_common(), columns=["activity", "count"])
    frame.to_csv(output / "dwelling_activity_counts.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(columns=["case_id", "truck_visit_minutes", "stack_to_unstack_hours"]).to_csv(
        output / "dwelling_case_metrics.csv", index=False, encoding="utf-8-sig"
    )


def _normalise(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for column in result.columns:
        if result[column].dtype == object:
            result[column] = result[column].fillna("").astype(str)
    if len(result.columns):
        try:
            result = result.sort_values(list(result.columns), kind="stable", na_position="last")
        except (TypeError, ValueError):
            pass
    return result.reset_index(drop=True)


def compare_csv(reference: Path, candidate: Path) -> dict[str, object]:
    old = _normalise(pd.read_csv(reference, encoding="utf-8-sig"))
    new = _normalise(pd.read_csv(candidate, encoding="utf-8-sig"))
    same_columns = list(old.columns) == list(new.columns)
    same_shape = old.shape == new.shape
    exact = False
    max_abs_difference = 0.0
    differing_cells = 0
    if same_columns and same_shape:
        comparisons = []
        for column in old.columns:
            if pd.api.types.is_numeric_dtype(old[column]) and pd.api.types.is_numeric_dtype(new[column]):
                left = old[column].to_numpy(dtype=float)
                right = new[column].to_numpy(dtype=float)
                valid = ~(np.isnan(left) & np.isnan(right))
                difference = np.abs(np.nan_to_num(left - right, nan=np.inf))
                comparisons.append(np.isclose(left, right, rtol=0.0, atol=1e-12, equal_nan=True))
                finite = difference[np.isfinite(difference)]
                if finite.size:
                    max_abs_difference = max(max_abs_difference, float(finite.max()))
            else:
                comparisons.append(old[column].to_numpy() == new[column].to_numpy())
        equal_matrix = np.column_stack(comparisons) if comparisons else np.ones((len(old), 0), dtype=bool)
        differing_cells = int((~equal_matrix).sum())
        exact = differing_cells == 0
    return {
        "file": reference.name,
        "reference_rows": len(old),
        "candidate_rows": len(new),
        "same_columns": same_columns,
        "same_shape": same_shape,
        "differing_cells": differing_cells if same_columns and same_shape else None,
        "max_abs_difference": max_abs_difference if same_columns and same_shape else None,
        "exact_match": exact,
        "reference_sha256": sha256(reference),
        "candidate_sha256": sha256(candidate),
    }


def scenario_delta(reference: Path, candidate: Path, label: str) -> pd.DataFrame:
    old = pd.read_csv(reference).set_index("scenario")
    new = pd.read_csv(candidate).set_index("scenario")
    rows = []
    for scenario in old.index.intersection(new.index):
        for metric in (
            "source_gate_missions",
            "demand_multiplier",
            "mean_requests",
            "missed_reservation_rate_from_virginia",
            "yard_utilization_from_houston",
            "capacity_disruption_probability",
        ):
            rows.append(
                {
                    "replay": label,
                    "scenario": scenario,
                    "metric": metric,
                    "archived": float(old.loc[scenario, metric]),
                    "replayed": float(new.loc[scenario, metric]),
                    "difference": float(new.loc[scenario, metric] - old.loc[scenario, metric]),
                }
            )
    return pd.DataFrame(rows)


def write_report(output: Path, archive_checks: pd.DataFrame, fresh_checks: pd.DataFrame, deltas: pd.DataFrame) -> Path:
    archived_exact = bool(archive_checks["exact_match"].all())
    fresh_exact = bool(fresh_checks["exact_match"].all())
    report = output / "E0_RAW_REPLAY_AUDIT.md"
    lines = [
        "# E0原始数据全链核验报告",
        "",
        "## 结论",
        "",
        f"- 论文日期快照全链精确匹配：{'是' if archived_exact else '否'}。",
        f"- 2026-07-17当前公开文件与论文日期快照精确匹配：{'是' if fresh_exact else '否'}。",
        "- SC Ports文件是滚动更新的公共仪表板导出，因此当前下载值发生变化不代表原校准错误；可复现论文数值需要保留带日期和哈希的论文期快照。",
        "- 当前下载重放用于检验校准对数据更新的敏感性，不能用来悄悄覆盖论文期校准输入。",
        "",
        "## 论文期留存快照与归档处理结果比较",
        "",
        archive_checks.to_markdown(index=False),
        "",
        "## 当前重新下载文件与归档处理结果比较",
        "",
        fresh_checks.to_markdown(index=False),
        "",
        "## 校准参数逐项差异",
        "",
        deltas.to_markdown(index=False),
        "",
        "## 可复现性表述",
        "",
        "论文期留存快照能够重建归档校准结果时，可以声称E0从归档原始快照到校准场景可复现。对于第三方重新下载，应说明滚动数据可能随下载日期变化；下载脚本、日期与SHA-256用于来源核验，而不是保证未来下载得到相同字节。",
    ]
    report.write_text("\n".join(lines), encoding="utf-8")
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Replay E0 from archival and freshly downloaded raw sources.")
    parser.add_argument("--output", type=Path, default=ROOT / "generated" / "e0_raw_replay")
    args = parser.parse_args()
    output = args.output
    output.mkdir(parents=True, exist_ok=True)
    processed = ROOT / "data" / "processed"

    archival_data = WORKSPACE / "data"
    archival_output = output / "archival_snapshot_replay"
    if not all((archival_output / name).exists() for name in OUTPUT_FILES):
        run_transform(
            sc_ports=archival_data / "sc_ports",
            virginia=archival_data / "port_virginia_weekly_metrics",
            houston=archival_data / "port_houston_terminal_reports",
            tas=archival_data / "mendeley_tas_tours",
            capacity=archival_data / "mendeley_capacity_management",
            dwelling=archival_data / "mendeley_dwelling_time",
            output=archival_output,
        )

    raw = ROOT / "data" / "raw"
    fresh_input = output / "fresh_extracted"
    mendeley = extract_mendeley(raw, fresh_input)
    fresh_output = output / "fresh_download_replay"
    run_transform(
        sc_ports=raw / "sc_ports",
        virginia=raw / "port_virginia_weekly_metrics",
        houston=raw / "port_houston_terminal_reports",
        tas=mendeley["mendeley_tas_tours"],
        capacity=mendeley["mendeley_capacity_management"],
        dwelling=mendeley["mendeley_dwelling_time"],
        output=fresh_output,
    )

    archive_checks = pd.DataFrame([compare_csv(processed / name, archival_output / name) for name in OUTPUT_FILES])
    fresh_checks = pd.DataFrame([compare_csv(processed / name, fresh_output / name) for name in OUTPUT_FILES])
    archive_checks.to_csv(output / "archival_snapshot_comparison.csv", index=False, encoding="utf-8-sig")
    fresh_checks.to_csv(output / "fresh_download_comparison.csv", index=False, encoding="utf-8-sig")
    deltas = pd.concat(
        [
            scenario_delta(processed / "public_data_calibrated_scenarios.csv", archival_output / "public_data_calibrated_scenarios.csv", "archival_snapshot"),
            scenario_delta(processed / "public_data_calibrated_scenarios.csv", fresh_output / "public_data_calibrated_scenarios.csv", "fresh_download"),
        ],
        ignore_index=True,
    )
    deltas.to_csv(output / "scenario_parameter_differences.csv", index=False, encoding="utf-8-sig")
    metadata = {
        "archival_input": "user-supplied retained paper-date snapshot (2026-07-09)",
        "fresh_manifest": "data/raw/download_manifest.csv",
        "processed_reference": "data/processed",
        "archival_all_exact": bool(archive_checks["exact_match"].all()),
        "fresh_all_exact": bool(fresh_checks["exact_match"].all()),
    }
    (output / "run_metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    report = write_report(output, archive_checks, fresh_checks, deltas)
    print(report)
    return 0 if bool(archive_checks["exact_match"].all()) else 2


if __name__ == "__main__":
    raise SystemExit(main())
